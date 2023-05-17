import requests
from bs4 import BeautifulSoup

url = 'http://umsf.dp.ua/'
response = requests.get(url)
html_res = response.text

soup = BeautifulSoup(html_res, 'html.parser')

inputs = soup.find_all()
input_list = []
for input_tag in inputs:
    input_list.append(input_tag)

import joblib
clf = joblib.load('model.pkl')

class_list = []
def classify_tags(html_res, clf, input_list, class_list):
    for item in input_list:
        bn = item
        bn=str(item)
        test_objects = [bn]
        X_test = vectorizer.transform(test_objects)
        predicted_classes = clf.predict(X_test)
        for i, obj in enumerate(test_objects):
            class_list.append(predicted_classes[i])
        test_objects.clear()
classify_tags(html_res, clf, input_list, class_list)

class_list = list(set(class_list))

import openpyxl
from openpyxl.styles import Font, Color, PatternFill, Alignment

def write_to_excel(checklist, class_list, url):

    wb = openpyxl.Workbook() #Створення нового документу
    sheet = wb.active #Активація листу

    sheet.column_dimensions['B'].width = 50 #Для ствопця В
    sheet.row_dimensions[1].height = 40
    cell_B1 = sheet["B1"]
    cell_B1.font = Font(size=12, color='FF0000')
    cell_B1.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    cell_B1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_range = sheet['C1:F1'] #Для C1-F1
    for row in cell_range:
        for cell in row:
            cell.font = Font(size=12, color='FF0000')
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 25
    sheet.column_dimensions['E'].width = 25
    sheet.column_dimensions['F'].width = 25
    sheet.column_dimensions['G'].width = 30
    sheet.column_dimensions['H'].width = 30

    cell_G1 = sheet["G1"]
    cell_G1.font = Font(size=12, color='FF0000')
    cell_G1.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    cell_G1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_H1 = sheet["H1"]
    cell_H1.font = Font(size=12, color='FF0000')
    cell_H1.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    cell_H1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    b1_text = "Сайт: " + url
    b1_text = str(b1_text)
    sheet["B1"] = b1_text
    sheet["C1"] = "Firefox 112.0.2"
    sheet["D1"] = "Google Chrome 100.0.4896.75"
    sheet["E1"] = "Opera 75.0.3969.171"
    sheet["F1"] = "Safari 14.1.2"
    sheet["G1"] = "Посилання на баг-репорт"
    sheet["H1"] = "Примітка"

    class_list.insert(0, "Дизайн")
    class_list.insert(0, "Верстка")
    class_list.insert(0, "Текст")
    class_list.insert(0, "Фавікон")

    sheet["C2"] = "Пройдено"
    sheet["D2"] = "Пройдено"
    sheet["E2"] = "Пропущено"
    sheet["F2"] = "Провалено"
    cell_C2 = sheet["C2"]
    cell_C2.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    cell_C2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_D2 = sheet["D2"]
    cell_D2.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    cell_D2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_E2 = sheet["E2"]
    cell_E2.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
    cell_E2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_F2 = sheet["F2"]
    cell_F2.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    cell_F2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    start_row = 2    
    for item in class_list:
        if item == "Дизайн":
            sheet.cell(row=start_row, column=1).value = "Дизайн"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[45:47]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1
        elif item == "Верстка":
            sheet.cell(row=start_row, column=1).value = "Верстка"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[48:54]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1
        elif item == "Текст":
            sheet.cell(row=start_row, column=1).value = "Текст"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[41:44]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1
        elif item == "Пошук":
            sheet.cell(row=start_row, column=1).value = "Пошук"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[0:6]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1
        elif item == "Кнопка":
            sheet.cell(row=start_row, column=1).value = "Кнопки"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[7:10]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1
        elif item == "Посилання":
            sheet.cell(row=start_row, column=1).value = "Посилання"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[11:13]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1
        elif item == "Зображення":
            sheet.cell(row=start_row, column=1).value = "Зображення"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[14:15]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1
        elif item == "Меню навігації":
            sheet.cell(row=start_row, column=1).value = "Меню навігації"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[16:19]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1
        elif item == "Підвал сайту":
            sheet.cell(row=start_row, column=1).value = "Підвал сайту"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[20:21]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1
        elif item == "Логотип":
            sheet.cell(row=start_row, column=1).value = "Логотип"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[22:25]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1
        elif item == "Слайдер":
            sheet.cell(row=start_row, column=1).value = "Слайдер"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[26:29]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1
        elif item == "Посилання на соцмережі (Фейсбук)":
            sheet.cell(row=start_row, column=1).value = "Посилання на соцмережі (Фейсбук)"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[30:32]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1
        elif item == "Посилання на соцмережі (Інстаграм)":
            sheet.cell(row=start_row, column=1).value = "Посилання на соцмережі (Інстаграм)"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[30:32]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1
        elif item == "Посилання на соцмережі (Твіттер)":
            sheet.cell(row=start_row, column=1).value = "Посилання на соцмережі (Твіттер)"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[30:32]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1  
        elif item == "Посилання на соцмережі (Ютуб)":
            sheet.cell(row=start_row, column=1).value = "Посилання на соцмережі (Ютуб)"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[30:32]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1
        elif item == "Скрипт":
            sheet.cell(row=start_row, column=1).value = "Скрипти"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[33:35]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1               
        elif item == "Копірайт":
            sheet.cell(row=start_row, column=1).value = "Копірайт"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[36:37]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1
        elif item == "Завантаження додатку (Гугл)":
            sheet.cell(row=start_row, column=1).value = "Завантаження додатку (Гугл)"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[38:40]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1
        elif item == "Завантаження додатку (Еппл)":
            sheet.cell(row=start_row, column=1).value = "Завантаження додатку (Еппл)"
            start_row += 1
            for col_num in range(1, 3):
                cell = sheet.cell(row=start_row-1, column=col_num)
                cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
            with open('content.txt') as f:
                lines = f.readlines()[38:40]
                print(lines)
                for line in lines:
                    sheet.cell(row=start_row, column=2).value = line.strip()
                    cell.alignment = Alignment(wrap_text=True)
                    start_row += 1

    wb.save('checklist.xlsx')

write_to_excel("checklist.xlsx", class_list, url)